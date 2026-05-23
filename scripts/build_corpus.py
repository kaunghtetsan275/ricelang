"""Build a fastText-format training corpus from kaunghtetsan275/corpus.

Consolidates the heterogeneous source files (already-labeled .txt, raw
single-blob .txt, and .xlsx exports) into ``data/train.txt`` and
``data/valid.txt`` ready for ``scripts/train_detector.py``.

Corpus file convention
----------------------
``<iso-639-3>_<source>.<ext>`` — language code prefix, source as postfix.
The Burmese files are tagged ``mya`` even though the model emits ``uni``
(Unicode) vs ``zg`` (Zawgyi) labels; the model's ``zg`` examples are
synthesized at corpus-build time from the ``mya`` Unicode text.

Sources handled
---------------
All ``*.txt`` files are read in fastText format (``__label__X text``).
- ``cnh_jsw.txt``       — label ``cnh``
- ``ksw_werribee.txt``  — label ``ksw``
- ``ksw_jsw.txt``       — label ``ksw``
- ``eky_kayahli.txt``   — label ``eky``
- ``eky_youversion.txt``— label ``eky`` (scraped Eastern Kayah NT)
- ``mya_jsw.txt``       — label ``mya`` (Burmese Unicode)
- ``mya_mmtimes.xlsx``  — Headline + Paragraph columns, label ``mya``
- (skipped) ``shn_shannews.xlsx`` — dirty training data per the project README
- (skipped) Mon         — no data available in this corpus

Optional ``zgi`` synthesis
--------------------------
The detector distinguishes Burmese Unicode (``mya``) from Zawgyi
(``zgi``). The corpus has no Zawgyi text, so by default we synthesize a
matched set by running ``pyidaungsu.cvt2zg`` over every ``mya`` example
(plus a sampled portion to keep classes balanced). Disable with
``--no-synthesize-zg``.
"""

from __future__ import annotations

import argparse
import random
import sys
from pathlib import Path

import openpyxl

from pyidaungsu.convert import cvt2zg

LABEL_PREFIX = "__label__"
MIN_LEN = 8  # drop trivially-short fragments


def _clean(text: str) -> str:
    """Collapse whitespace; fastText needs one example per line."""
    return " ".join(text.split())


def _emit(label: str, text: str) -> tuple[str, str] | None:
    text = _clean(text)
    if len(text) < MIN_LEN:
        return None
    return label, text


def _read_labeled_file(path: Path) -> list[tuple[str, str]]:
    """Read a file that's already in fastText format (``__label__x  text``)."""
    out: list[tuple[str, str]] = []
    with path.open(encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line.startswith(LABEL_PREFIX):
                continue
            head, _, text = line.partition("\t")
            if not text:
                head, _, text = line.partition(" ")
            label = head[len(LABEL_PREFIX):]
            example = _emit(label, text)
            if example:
                out.append(example)
    return out


def _read_xlsx(path: Path, label: str, columns: list[str]) -> list[tuple[str, str]]:
    """Extract text from named columns of a single-sheet xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idxs = [header.index(c) for c in columns]
    out: list[tuple[str, str]] = []
    for row in rows:
        for i in idxs:
            cell = row[i] if i < len(row) else None
            if not cell:
                continue
            example = _emit(label, str(cell))
            if example:
                out.append(example)
    return out


def collect(corpus_dir: Path, synthesize_zg: bool, zg_ratio: float) -> list[tuple[str, str]]:
    examples: list[tuple[str, str]] = []

    # Pre-labeled fastText files. Filename language code matches the
    # label (e.g. mya_*.txt -> __label__mya); zgi (Zawgyi-encoded
    # Burmese) is synthesized from mya text below.
    for name in (
        "cnh_jsw.txt",
        "ksw_werribee.txt",
        "ksw_jsw.txt",
        "eky_kayahli.txt",
        "eky_youversion.txt",
        "mya_jsw.txt",
    ):
        path = corpus_dir / name
        if path.exists():
            examples += _read_labeled_file(path)

    # xlsx exports
    examples += _read_xlsx(corpus_dir / "mya_mmtimes.xlsx", "mya",
                           columns=["Headline", "Paragraph"])

    # shn_shannews.xlsx is intentionally skipped (dirty training data).

    if synthesize_zg:
        mya_examples = [t for lbl, t in examples if lbl == "mya"]
        # Sample to control class balance; default 1.0 = match mya count.
        target_n = int(len(mya_examples) * zg_ratio)
        sampled = random.sample(mya_examples, min(target_n, len(mya_examples)))
        for text in sampled:
            zg = cvt2zg(text)
            if zg and zg != text:
                examples.append(("zgi", zg))

    return examples


def write_fasttext(path: Path, examples: list[tuple[str, str]]) -> None:
    with path.open("w", encoding="utf-8") as fh:
        for label, text in examples:
            fh.write(f"{LABEL_PREFIX}{label} {text}\n")


def summarize(examples: list[tuple[str, str]]) -> str:
    counts: dict[str, int] = {}
    for label, _ in examples:
        counts[label] = counts.get(label, 0) + 1
    width = max(len(l) for l in counts)
    lines = [f"  {label.ljust(width)}  {n:>7,}" for label, n in sorted(counts.items())]
    lines.append(f"  {'total'.ljust(width)}  {len(examples):>7,}")
    return "\n".join(lines)


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--corpus", default="../corpus/data",
                   help="path to corpus data directory (default: ../corpus/data)")
    p.add_argument("--out", default="data", help="output directory (default: data/)")
    p.add_argument("--valid-fraction", type=float, default=0.1)
    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--no-synthesize-zg", dest="synthesize_zg", action="store_false",
                   help="don't synthesize zg examples from uni text via cvt2zg")
    p.add_argument("--zg-ratio", type=float, default=1.0,
                   help="number of synthetic zg examples relative to uni count (default 1.0)")
    args = p.parse_args(argv)

    random.seed(args.seed)
    corpus_dir = Path(args.corpus).resolve()
    out_dir = Path(args.out).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"[corpus] reading from {corpus_dir}")
    examples = collect(corpus_dir, args.synthesize_zg, args.zg_ratio)
    print(f"[corpus] collected:")
    print(summarize(examples))

    random.shuffle(examples)
    split = int(len(examples) * (1 - args.valid_fraction))
    train, valid = examples[:split], examples[split:]

    train_path = out_dir / "train.txt"
    valid_path = out_dir / "valid.txt"
    write_fasttext(train_path, train)
    write_fasttext(valid_path, valid)
    print(f"[write] train -> {train_path}  ({len(train):,} lines)")
    print(f"[write] valid -> {valid_path}  ({len(valid):,} lines)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
